import customtkinter as ctk
import openpyxl as xl


class App(ctk.CTk):
    def __init__(self, TITLE, GEOMETRY):
        super().__init__()
        self.title(TITLE)
        self.geometry(GEOMETRY)
        self.resizable(False, False)

        self.create_grid()

        self.task_frame = TaskFrame(self)
        self.main_frame = MainFrame(self, self.task_frame)

        self.main_frame.grid(row = 0, column = 0, sticky = "nswe")
        self.task_frame.grid(row = 1, column = 0, sticky = "nswe")

        self.mainloop()


    def create_grid(self):
        self.columnconfigure(0, weight = 1)
        self.rowconfigure(0, weight = 1)
        self.rowconfigure(1, weight = 100)



class MainFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent, fg_color = "mediumorchid3", corner_radius = 0)

        self.controller = controller

        self.create_variables()
        self.create_fonts()
        self.create_widgets()
        self.create_grid()
        self.display_widgets()


    def create_variables(self):
        self.task_var = ctk.StringVar()


    def add_task(self):

        task = self.task_var.get().strip()

        if task == "":
            return

        self.wb = xl.load_workbook("tasks.xlsx")
        self.ws = self.wb.active

        self.ws.append([task])
        self.wb.save("tasks.xlsx")

        self.task_entry.delete(0, "end")

        # refresh task list
        for widget in self.controller.text_frame.winfo_children():
            widget.destroy()

        self.controller.create_widgets()


    def create_fonts(self):
        self.title_font = ctk.CTkFont(family = "helvetica", size = 40, weight = "bold")
        self.subtitle_font = ctk.CTkFont(family = "helvetica", size = 30, weight = "bold")
        self.normal_font = ctk.CTkFont(family = "helvetica", size = 20, weight = "bold")


    def create_widgets(self):

        self.app_title = ctk.CTkLabel(master = self,
                                      text_color = "plum1",
                                      text = "To-Do List",
                                      font = self.title_font,
                                      pady = 20)

        self.task_entry = ctk.CTkEntry(master = self,
                                       text_color = "plum1",
                                       fg_color = "transparent",
                                       corner_radius = 10,
                                       border_width = 5,
                                       border_color = "mediumorchid2",
                                       width = 300,
                                       height = 50,
                                       font = self.subtitle_font,
                                       textvariable = self.task_var)

        self.add_button = ctk.CTkButton(master = self,
                                        text = "Add Task",
                                        command = self.add_task,
                                        width = 200,
                                        height = 50,
                                        fg_color = "mediumorchid2",
                                        font = self.subtitle_font,
                                        hover_color = "mediumorchid4")


    def create_grid(self):
        self.columnconfigure(0, weight = 1)
        self.rowconfigure(0, weight = 1)
        self.rowconfigure(1, weight = 1)


    def display_widgets(self):
        self.app_title.grid(row = 0, column = 0)

        self.task_entry.grid(row = 1, column = 0,
                             sticky = "w",
                             padx = 45,
                             pady = 20)

        self.add_button.grid(row = 1, column = 0,
                             sticky = "e",
                             padx = 45,
                             pady = 20)



class TaskFrame(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color = "mediumorchid2", corner_radius = 0)

        self.create_workbook()
        self.create_frame()
        self.create_font()
        self.create_widgets()


    def create_workbook(self):

        try:
            self.wb = xl.load_workbook("tasks.xlsx")
            self.ws = self.wb.active

        except:
            self.wb = xl.Workbook()
            self.ws = self.wb.active
            self.wb.save("tasks.xlsx")


    def create_frame(self):

        self.text_frame = ctk.CTkScrollableFrame(master = self,
                                                 corner_radius = 10,
                                                 border_width = 5,
                                                 border_color = "mediumorchid3",
                                                 fg_color = "transparent")

        self.text_frame.pack(fill = "both",
                             expand = True,
                             padx = 10,
                             pady = 10)

        self.text_frame.columnconfigure(0, weight = 1)
        self.text_frame.columnconfigure(1, weight = 0)


    def create_font(self):
        self.normal_font = ctk.CTkFont(family = "helvetica", size = 20, weight = "bold")


    def delete_task(self, row):

        self.wb = xl.load_workbook("tasks.xlsx")
        self.ws = self.wb.active

        tasks = []

        for r in range(1, self.ws.max_row + 1):

            value = self.ws[f"A{r}"].value

            if value != None and r != row:
                tasks.append(value)

        self.wb = xl.Workbook()
        self.ws = self.wb.active

        for task in tasks:
            self.ws.append([task])

        self.wb.save("tasks.xlsx")

        for widget in self.text_frame.winfo_children():
            widget.destroy()

        self.create_widgets()


    def create_widgets(self):

        self.wb = xl.load_workbook("tasks.xlsx")
        self.ws = self.wb.active

        row_index = 0

        for excel_row in range(1, self.ws.max_row + 1):

            task_text = self.ws[f"A{excel_row}"].value

            if task_text == None:
                continue

            label = ctk.CTkLabel(master = self.text_frame,
                                 text_color = "plum1",
                                 text = task_text,
                                 font = self.normal_font,
                                 wraplength = 450,
                                 justify = "left")

            delete_button = ctk.CTkButton(master = self.text_frame,
                                          text = "Delete",
                                          text_color = "plum1",
                                          font = self.normal_font,
                                          width = 30,
                                          height = 30,
                                          command = lambda r = excel_row: self.delete_task(r),
                                          border_color = "mediumorchid3",
                                          fg_color = "transparent",
                                          corner_radius = 5,
                                          border_width = 5,
                                          hover_color = "mediumorchid2",
                                          border_spacing = 10)

            label.grid(row = row_index,
                       column = 0,
                       sticky = "w",
                       padx = 10,
                       pady = 5)

            delete_button.grid(row = row_index,
                               column = 1,
                               sticky = "e",
                               padx = 10,
                               pady = 5)

            row_index += 1



if __name__ == "__main__":
    App("To-Do List", "600x600")